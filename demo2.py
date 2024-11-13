import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Load the Excel file into a DataFrame
df = pd.read_excel('output.xlsx')

# Define the folder path
folder_path = 'C:\\Folder\\'

# Concatenate the folder path with the filename column
df['Full Path'] = folder_path + df['Name'].astype(str)

# Save the DataFrame to a temporary Excel file
df.to_excel('temp_file.xlsx', index=False)

# Load the temporary Excel file with openpyxl
wb = load_workbook('temp_file.xlsx')
ws = wb.active

# Add hyperlinks to the 'Full Path' column
for row in range(2, ws.max_row + 1):  # Start from the second row
    cell = ws.cell(row=row, column=ws.max_column)
    cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
    cell.font = Font(color="0000FF", underline="single")

# Save the updated Excel file
wb.save('updated_file_with_hyperlinks.xlsx')

print("New column with concatenated paths and hyperlinks added and saved to updated_file_with_hyperlinks.xlsx")
